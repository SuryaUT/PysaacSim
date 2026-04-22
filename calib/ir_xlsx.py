"""Fit per-side IR firmware constants from IR_Calib.xlsx.

Firmware formula (RTOS_Labs_common/IRDistance.c):
    d_mm = a / (adc + b) + c                 (adc >= adc_threshold)

The workbook as of 2026-04-22 has a single sheet with columns
    (inch, IR_Left_ADC, IR_RIGHT_ADC)
— no separate TFLuna sheet. We convert inch→mm and fit `(a, b, c)` per side
via `scipy.optimize.curve_fit`. `adc_threshold` is the lowest ADC value whose
fitted distance is ≤ 305 mm; that matches the firmware's out-of-range logic.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

import numpy as np
from scipy.optimize import curve_fit

import openpyxl


INCH_TO_MM = 25.4
SAT_MM = 305.0   # firmware out-of-range sentinel


@dataclass
class IRSideFit:
    side: str            # 'left' or 'right'
    a: float
    b: float
    c: float
    adc_threshold: int
    rmse_mm: float
    n_points: int


@dataclass
class IRXlsxFit:
    left: IRSideFit
    right: IRSideFit
    inches: np.ndarray
    adc_left: np.ndarray
    adc_right: np.ndarray
    mm: np.ndarray


def _read_sheet(path: Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True) if r and r[0] is not None]
    header = [str(c).strip().lower() for c in rows[0]]
    data = rows[1:]
    # find columns — tolerate case and underscore variations
    def col(name_substr: str) -> int:
        for i, h in enumerate(header):
            if name_substr in h:
                return i
        raise KeyError(f"column matching {name_substr!r} not found in {header}")
    i_inch  = col("inch")
    i_left  = col("left")
    i_right = col("right")
    inches = np.array([float(r[i_inch])  for r in data])
    al     = np.array([float(r[i_left])  for r in data])
    ar     = np.array([float(r[i_right]) for r in data])
    return inches, al, ar


def _fit_side(adc: np.ndarray, mm: np.ndarray, side: str) -> IRSideFit:
    def model(x, a, b, c):
        return a / (x + b) + c
    adc_min = float(adc.min())
    # Try several initial guesses; nonlinear fit is sensitive to p0 and the
    # firmware's old constants aren't a good anchor once the xlsx changes.
    p0_candidates = [
        (100_000.0,    0.0,  0.0),
        (100_000.0, -500.0,  0.0),
        (50_000.0,  -200.0, 20.0),
        (137_932.0, -859.0, 32.0),   # firmware Left
        (52_850.0, -1239.0, 69.0),   # firmware Right
        (268_130.0,  159.0,  0.0),   # firmware generic (unscaled)
    ]
    best: Optional[tuple[float, tuple[float, float, float]]] = None
    for p0 in p0_candidates:
        try:
            popt, _ = curve_fit(model, adc, mm, p0=p0, maxfev=40_000)
            a, b, c = map(float, popt)
            if not (np.isfinite(a) and np.isfinite(b) and np.isfinite(c)):
                continue
            # Reject nonphysical fits (`a` should be positive so d decreases
            # with adc in the operating range).
            if a <= 0:
                continue
            # Reject singular fits (denominator crosses 0 inside the adc range).
            if (adc_min + b) * (adc.max() + b) <= 0:
                continue
            resid = mm - (a / (adc + b) + c)
            rmse = float(np.sqrt(np.mean(resid * resid)))
            if best is None or rmse < best[0]:
                best = (rmse, (a, b, c))
        except Exception:
            continue
    if best is None:
        raise RuntimeError(
            f"IR fit did not converge for side={side!r}; data n={adc.size}"
        )
    rmse, (a, b, c) = best
    # adc_threshold: where the fitted distance crosses the firmware's 305 mm
    # sentinel. Rounded *up* to the nearest integer so adc < threshold → OOR.
    denom = (SAT_MM - c)
    if denom > 0:
        adc_at_sat = int(np.ceil((a / denom) - b))
    else:
        adc_at_sat = int(np.floor(adc_min))
    # Never threshold above the observed adc range.
    adc_at_sat = max(adc_at_sat, int(np.floor(adc_min)))
    return IRSideFit(
        side=side, a=a, b=b, c=c, adc_threshold=adc_at_sat,
        rmse_mm=rmse, n_points=int(adc.size),
    )


def fit_xlsx(path: Union[str, Path]) -> IRXlsxFit:
    p = Path(path)
    inches, adc_l, adc_r = _read_sheet(p)
    mm = inches * INCH_TO_MM
    left = _fit_side(adc_l, mm, "left")
    right = _fit_side(adc_r, mm, "right")
    return IRXlsxFit(
        left=left, right=right,
        inches=inches, adc_left=adc_l, adc_right=adc_r, mm=mm,
    )
