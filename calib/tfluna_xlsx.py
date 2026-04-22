"""Fit per-lidar TFLuna calibration, if a suitable sheet exists.

The current `IR_Calib.xlsx` (as of 2026-04-22) has **no** TFLuna sheet — only
a single sheet with (inch, IR_Left_ADC, IR_Right_ADC). This module therefore
ships as a no-op: `fit_xlsx` returns `None` when no TFLuna columns are found,
and the caller (report.py / scripts/calibrate_from_log.py) simply leaves the
lidar calibration at its defaults.

If a future workbook adds sheets with columns like
(true_mm, tf_left_mm, tf_middle_mm, tf_right_mm), this module auto-detects and
fits `reported = scale * true + bias` + residual std per lidar.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

import numpy as np

import openpyxl


@dataclass
class TFLunaSideFit:
    id: str
    scale: float
    bias_cm: float
    noise_std_cm: float
    n_points: int


@dataclass
class TFLunaXlsxFit:
    per_lidar: dict[str, TFLunaSideFit]


def _find_tfluna_sheet(wb) -> Optional[tuple[str, list[tuple]]]:
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [r for r in ws.iter_rows(values_only=True) if r and r[0] is not None]
        if not rows:
            continue
        header = [str(c).strip().lower() for c in rows[0] if c is not None]
        # Heuristic: at least one TFLuna-like column name
        if any("tf" in h or "tfluna" in h or "lidar" in h for h in header):
            return name, rows
    return None


def fit_xlsx(path: Union[str, Path]) -> Optional[TFLunaXlsxFit]:
    p = Path(path)
    wb = openpyxl.load_workbook(p, data_only=True)
    found = _find_tfluna_sheet(wb)
    if found is None:
        return None
    _, rows = found
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    data = rows[1:]
    # Column index helpers.
    def opt_col(substr: str) -> Optional[int]:
        for i, h in enumerate(header):
            if substr in h:
                return i
        return None
    i_true = opt_col("true")
    if i_true is None:
        # If no "true_mm" column is present, we can't fit. Just bail.
        return None
    true_mm = np.array([float(r[i_true]) for r in data])
    fits: dict[str, TFLunaSideFit] = {}
    for label in ("tf_left", "tf_middle", "tf_right"):
        i = opt_col(label)
        if i is None:
            continue
        reported = np.array([float(r[i]) for r in data])
        # Linear LS fit: reported = scale*true + bias (in same units as sheet).
        A = np.column_stack([true_mm, np.ones_like(true_mm)])
        (scale, bias), *_ = np.linalg.lstsq(A, reported, rcond=None)
        resid = reported - (scale * true_mm + bias)
        fits[label] = TFLunaSideFit(
            id=label,
            scale=float(scale),
            bias_cm=float(bias) / 10.0,
            noise_std_cm=float(np.std(resid)) / 10.0,
            n_points=int(reported.size),
        )
    if not fits:
        return None
    return TFLunaXlsxFit(per_lidar=fits)
